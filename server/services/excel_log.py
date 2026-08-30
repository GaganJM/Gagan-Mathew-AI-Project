import threading
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import config

HEADERS = [
    "Reference Number",
    "Company Name",
    "Submission Date/Time",
    "Assigned To",
    "Submission ID",
    "Extraction Confidence",
    "Local Folder",
    "Google Drive Folder",
    "Flags",
    "Contact Email(s)",
    "Contact Number(s)",
]

REFERENCE_PREFIX = "WBG-26"
REFERENCE_DIGITS = 3

PROFILE_CHANGE_SHEET_NAME = "Profile Change"
PROFILE_CHANGE_HEADERS = [
    "Date of Submission",
    "Company Name",
    "CIF Number",
    "Account Number",
    "Reference Number",
    "Type of Changes",
    "Local Folder",
    "Google Drive Folder",
    "Contact Email(s)",
    "Contact Number(s)",
]
PROFILE_CHANGE_REFERENCE_PREFIX = "PC26"
PROFILE_CHANGE_REFERENCE_DIGITS = 3

HYPERLINK_FONT = Font(color="0563C1", underline="single")

_lock = threading.Lock()


def _local_folder_uri(folder_path):
    return "file:///" + str(Path(folder_path)).replace("\\", "/")


def _next_reference_number(ws, prefix, digits):
    existing_data_rows = max(ws.max_row - 1, 0)  # exclude header row
    return f"{prefix}{existing_data_rows + 1:0{digits}d}"


def _load_or_create_workbook(excel_path):
    if excel_path.exists():
        return load_workbook(excel_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Submissions"
    ws.append(HEADERS)
    ws.freeze_panes = "A2"
    return wb


def _get_or_create_sheet(wb, title, headers):
    if title in wb.sheetnames:
        return wb[title]
    ws = wb.create_sheet(title)
    ws.append(headers)
    ws.freeze_panes = "A2"
    return ws


def append_submission_row(
    company_name,
    submitted_at_iso,
    assigned_to,
    submission_id,
    confidence,
    folder_path,
    drive_link,
    flags,
):
    """Appends a row to the persistent submissions tracker and returns the
    newly assigned reference number (e.g. 'WBG-26003')."""
    excel_path = config.EXCEL_LOG_PATH
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    with _lock:
        wb = _load_or_create_workbook(excel_path)
        ws = wb["Submissions"] if "Submissions" in wb.sheetnames else wb.active

        reference_number = _next_reference_number(ws, REFERENCE_PREFIX, REFERENCE_DIGITS)

        ws.append([
            reference_number, company_name, submitted_at_iso, assigned_to,
            submission_id, confidence, "", "", flags,
        ])
        row_idx = ws.max_row
        last_col = get_column_letter(len(HEADERS))
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

        local_cell = ws.cell(row=row_idx, column=7, value="Open local folder")
        local_cell.hyperlink = _local_folder_uri(folder_path)
        local_cell.font = HYPERLINK_FONT

        if drive_link:
            drive_cell = ws.cell(row=row_idx, column=8, value="Open in Google Drive")
            drive_cell.hyperlink = drive_link
            drive_cell.font = HYPERLINK_FONT
        else:
            ws.cell(row=row_idx, column=8, value="Not available")

        wb.save(excel_path)

    return reference_number


def append_profile_change_row(
    company_name,
    submitted_at_iso,
    cif_number,
    account_number,
    change_types,
    folder_path,
    drive_link,
):
    """Appends a row to the 'Profile Change' worksheet (created on first use)
    and returns the newly assigned reference number (e.g. 'PC26003'). Exactly
    one of cif_number/account_number should be a non-empty string; the other
    should be an empty string."""
    excel_path = config.EXCEL_LOG_PATH
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    with _lock:
        wb = _load_or_create_workbook(excel_path)
        ws = _get_or_create_sheet(wb, PROFILE_CHANGE_SHEET_NAME, PROFILE_CHANGE_HEADERS)

        reference_number = _next_reference_number(ws, PROFILE_CHANGE_REFERENCE_PREFIX, PROFILE_CHANGE_REFERENCE_DIGITS)

        ws.append([
            submitted_at_iso, company_name, cif_number, account_number,
            reference_number, "; ".join(change_types), "", "", "", "",
        ])
        row_idx = ws.max_row
        last_col = get_column_letter(len(PROFILE_CHANGE_HEADERS))
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

        local_col = PROFILE_CHANGE_HEADERS.index("Local Folder") + 1
        drive_col = PROFILE_CHANGE_HEADERS.index("Google Drive Folder") + 1

        local_cell = ws.cell(row=row_idx, column=local_col, value="Open local folder")
        local_cell.hyperlink = _local_folder_uri(folder_path)
        local_cell.font = HYPERLINK_FONT

        if drive_link:
            drive_cell = ws.cell(row=row_idx, column=drive_col, value="Open in Google Drive")
            drive_cell.hyperlink = drive_link
            drive_cell.font = HYPERLINK_FONT
        else:
            ws.cell(row=row_idx, column=drive_col, value="Not available")

        wb.save(excel_path)

    return reference_number


def update_contact_info(reference_number, emails, phones):
    """Finds the row matching reference_number (searching every worksheet
    that has Reference Number / Contact Email(s) / Contact Number(s) columns)
    and fills in the contact email(s) and phone number(s). Returns True if a
    matching row was found and updated, False otherwise."""
    excel_path = config.EXCEL_LOG_PATH
    if not excel_path.exists():
        return False

    with _lock:
        wb = load_workbook(excel_path)
        found = False

        for ws in wb.worksheets:
            headers = [c.value for c in ws[1]] if ws.max_row >= 1 else []
            if not {"Reference Number", "Contact Email(s)", "Contact Number(s)"}.issubset(headers):
                continue

            ref_col = headers.index("Reference Number") + 1
            email_col = headers.index("Contact Email(s)") + 1
            phone_col = headers.index("Contact Number(s)") + 1

            for row in ws.iter_rows(min_row=2):
                if row[ref_col - 1].value == reference_number:
                    row_idx = row[0].row
                    ws.cell(row=row_idx, column=email_col, value="; ".join(emails))
                    ws.cell(row=row_idx, column=phone_col, value="; ".join(phones))
                    found = True
                    break

            if found:
                break

        if found:
            wb.save(excel_path)

    return found
